from openpyxl import load_workbook
from openpyxl import Workbook

"""
This parses Mortpak output document and produces a CSV file with key data
C Gould
github(s): roly97, conorg000
"""

def get_data(workbook):
    """
    Goes through excel workbook and extracts data which comes under the
    heading "$ Population by single year of age"
    :param workbook: name of Excel workbook containing Mortpak output
    :return: list of lists, each one containing that year's data
    """
    try:
        wb1 = load_workbook(filename=workbook)
    except FileError:
        print("Couldn't find that file!")
        sys.exit()
    sheet = wb1.worksheets[0]
    data = []
    count = 0
    # Make a list of lists
    # Each list has that year's data
    for row in sheet.iter_rows(min_row=1, min_col=1, max_col=1):
        for cell in row:
            if "$ POPULATION BY SINGLE YEAR OF AGE" in str(cell.value):
                # Adde a new list to data
                data.append([])
                # Get 46 rows of data underneath the title
                end = 'A' + str(int(str(cell.coordinate)[1:]) + 46)
                for val in sheet[cell.coordinate: end]:
                    for obj in val:
                        data[count].append(obj.value)
                count += 1
    #print(data)
    return data

# Some data is stuck in elements with other data
# Split into two lists, then merge
def wrangle(data):
    """
    Rearranges some of the data, as some lists contain more than one age group per element
    :param data: list of lists, each one containing that year's data
    :return: list of lists, each one containing that year's data
    """
    new_data = []
    for year in data:
        part_one = []
        part_two = []
        part_one.append(year[:1])
        for vals in year[2:]:
            space = vals.split(' ')
            if len(space) > 4:
                part_one.append(space[:4])
                part_two.append(space[4:])
            else:
                part_one.append(space[:4])
        new_data.append(part_one + part_two)
    #print(new_data)
    return new_data

def to_csv(new_data):
    """
    Gets list of data, writes to a csv file
    :param new_data: list of lists, each one containing that year's data
    :return: csv file
    """
    fhand = open("mortpak_results.csv", "w")
    for year in new_data:
        fhand.write(''.join(year[0]) + '\n')
        for entry in year[1:]:
            fhand.write(','.join(entry) + '\n')
    fhand.close()

# Get name of excel workbook containing data
input = input('Enter name of excel workbook: ')
# Get data from workbook
data = get_data(input + '.xlsx')
# Edit the list of data
new_data = wrangle(data)
# Save to csv
to_csv(new_data)
